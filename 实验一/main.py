"""
实现excel转json
以及excel转bin
"""

import openpyxl
import json
import io
import struct
import os
from datetime import datetime


class DateEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.strftime("%Y-%m-%d")
        else:
            return json.JSONEncoder.default(self, obj)


# excel表格转json文件
def excel_to_json(excel_file, json_file_name):
    # 加载工作薄
    book = openpyxl.load_workbook(excel_file)
    # 获取sheet页
    sheet = book["Sheet1"]
    # 行数
    max_row = sheet.max_row
    # 列数
    max_column = sheet.max_column
    print("max_row: %d, max_column: %d" % (max_row, max_column))
    # 结果，数组存储
    result = []
    heads = []
    # 解析表头
    for column in range(max_column):
        # 读取的话行列是从（1，1）开始
        heads.append(sheet.cell(1, column + 1).value)
    # 遍历每一行
    for row in range(max_row):
        if row == 0:
            continue
        one_line = {}
        for column in range(max_column):
            # 读取第二行开始每一个数据
            k = heads[column]
            cell = sheet.cell(row + 1, column + 1)
            value = cell.value
            one_line[k] = value
        print(one_line)
        result.append(one_line)
    book.close()
    # 将json保存为文件
    save_json_file(result, json_file_name)

# 将json保存为文件
def save_json_file(jd, json_file_name):
    file = io.open(json_file_name, 'w', encoding='utf-8')
    # 把对象转化为json对象
    # indent: 参数根据数据格式缩进显示，读起来更加清晰
    # ensure_ascii = True：默认输出ASCII码，如果把这个该成False, 就可以输出中文。
    txt = json.dumps(jd, indent=2, ensure_ascii=False, cls=DateEncoder)
    file.write(txt)
    file.close()


# 将excel存储为二进制文件
def excel_to_bin(excel_file, bin_file_name):
    # 加载工作薄
    book = openpyxl.load_workbook(excel_file)
    # 获取sheet页
    sheet = book["Sheet1"]
    # 行数
    max_row = sheet.max_row
    # 列数
    max_column = sheet.max_column
    print("max_row: %d, max_column: %d" % (max_row, max_column))
    with open(bin_file_name, "wb") as f:
        for i in range(1, max_row):
            for j in range(1, max_column):
                if type(sheet.cell(i, j).value) is str: 
                    f.write(sheet.cell(i, j).value.encode())
                elif type(sheet.cell(i, j).value) is int: 
                    s = struct.pack("i", sheet.cell(i, j).value)
                    f.write(s)
                else:
                    f.write(sheet.cell(i, j).value)
    book.close()


# 返回文件大小
def return_size(filename):
    return f"{filename}文件的大小为{os.path.getsize(filename)}"


if '__main__' == __name__:

    excel_to_json(r'xlsx\sp_classes.xlsx', 'json\sp_classes.json')
    # excel_to_bin(r"xlsx\sp_classes.xlsx", "bin\sp_classes.bin")

    excel_to_json(r'xlsx\sp_course.xlsx', 'json\sp_course.json')
    # excel_to_bin(r"xlsx\sp_course.xlsx", "bin\sp_course.bin")

    excel_to_json(r'xlsx\sp_score_record.xlsx', 'json\sp_score_record.json')
    # excel_to_bin(r"xlsx\sp_score_record.xlsx", "bin\sp_score_record.bin")

    excel_to_json(r'xlsx\sp_teacher.xlsx', 'json\sp_teacher.json')
    # excel_to_bin(r"xlsx\sp_teacher.xlsx", "bin\sp_teacher.bin")

    excel_to_json(r'xlsx\sq_student.xlsx', 'json\sq_student.json')
    # excel_to_bin(r"xlsx\sq_student.xlsx", "bin\sq_student.bin")
    
    # 比较文件大小
    print(return_size(r'xlsx\sp_classes.xlsx'))
    print(return_size(r'json\sp_classes.json'))

    print(return_size(r'xlsx\sp_course.xlsx'))
    print(return_size(r'json\sp_course.json'))

    print(return_size(r'xlsx\sp_score_record.xlsx'))
    print(return_size(r'xlsx\sp_classes.xlsx'))

    print(return_size(r'xlsx\sp_teacher.xlsx'))
    print(return_size(r'json\sp_teacher.json'))

    print(return_size(r'xlsx\sq_student.xlsx'))
    print(return_size(r'json\sq_student.json'))
